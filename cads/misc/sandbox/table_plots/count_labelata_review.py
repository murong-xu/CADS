from openpyxl import load_workbook
import pandas as pd

def read_xlsx_to_dict(file_path):
    """
    Reads an XLSX file and returns a dictionary with Image ID as key,
    and a nested dictionary of structure reliability as value.
    """
    data = {}

    wb = load_workbook(filename=file_path, read_only=True)
    sheet = wb.active

    headers = [cell.value for cell in next(sheet.rows)]

    image_id_idx = headers.index("Image ID")
    label_idx = headers.index("Label")
    reliable_idx = headers.index("Relaible (Y/N)")
    comment_idx = headers.index('Comment')

    for row in sheet.iter_rows(min_row=2):  # skip header
        image_id = row[image_id_idx].value
        label = row[label_idx].value
        reliable = str(row[reliable_idx].value).strip().upper()

        # process reliability
        if reliable == 'NONE':
            comment = row[comment_idx].value
            if comment and ('not present' in comment or 'minimal part present' in comment):
                # limited field of view, ignore any potential FN
                is_reliable = -2
            elif comment and 'not labelled' in comment:
                # unreliable GT
                is_reliable = False
            else:
                is_reliable = True
        elif reliable == 'TBC':
            is_reliable = -2
        elif reliable == 'N':
            is_reliable = False
        elif reliable == 'Y':
            is_reliable = True
        else:
            # Default to True if reliability is not explicitly N, TBC, or specific NONE cases
            is_reliable = True


        if image_id not in data:
            data[image_id] = {}
        data[image_id][label] = is_reliable

    return data

def organize_by_reliability_and_structure(labelata_review_data):
    """
    Organizes the labelata review data into a dictionary structured by
    reliability status, then by structure, then a list of image IDs.
    """
    organized_data = {
        -2: {},  # Reliability status -2
        False: {},  # Reliability status False
        True: {}   # Reliability status True
    }

    for image_id, structures in labelata_review_data.items():
        for structure, reliability in structures.items():
            if reliability in organized_data:
                if structure not in organized_data[reliability]:
                    organized_data[reliability][structure] = []
                organized_data[reliability][structure].append(image_id)

    return organized_data

def calculate_reliability_counts(organized_data):
    """
    Calculates the count of image IDs for each structure within each reliability status.
    Returns a pandas DataFrame sorted by 'Reliability Status' and 'Image ID Count' descending.
    """
    counts_data = []
    for status, structures in organized_data.items():
        for structure, image_ids in structures.items():
            counts_data.append({
                "Reliability Status": status,
                "Structure": structure,
                "Image ID Count": len(image_ids)
            })
    
    df_counts = pd.DataFrame(counts_data)
    if not df_counts.empty:
        # Sort by Reliability Status and then by Image ID Count in descending order
        df_counts = df_counts.sort_values(
            by=["Reliability Status", "Image ID Count"],
            ascending=[True, False]
        ).reset_index(drop=True)
    return df_counts

def save_results_to_excel(organized_data, df_counts, output_filepath):
    """
    Saves the organized review data and reliability counts to an Excel file,
    with structures sorted by count within each reliability status.
    """
    with pd.ExcelWriter(output_filepath, engine='openpyxl') as writer:
        # Save counts data to a sheet (already sorted from calculate_reliability_counts)
        df_counts.to_excel(writer, sheet_name="Structure_Reliability_Counts", index=False)

        # Prepare and save organized review data to another sheet
        flattened_data = []
        for status, structures in organized_data.items():
            for structure, image_ids in structures.items():
                flattened_data.append({
                    "Reliability Status": status,
                    "Structure": structure,
                    "Image IDs": ", ".join(map(str, image_ids)),
                    "Image ID Count_temp": len(image_ids) # Temporary column for sorting
                })
        
        df_organized = pd.DataFrame(flattened_data)
        if not df_organized.empty:
            # Sort by Reliability Status and then by Image ID Count_temp in descending order
            df_organized = df_organized.sort_values(
                by=["Reliability Status", "Image ID Count_temp"],
                ascending=[True, False]
            ).drop(columns=["Image ID Count_temp"]).reset_index(drop=True) # Drop temp column before saving
        else:
            # Create an empty DataFrame with headers if no data
            empty_df = pd.DataFrame(columns=["Reliability Status", "Structure", "Image IDs"])
            df_organized = empty_df # Assign empty_df to df_organized


        df_organized.to_excel(writer, sheet_name="Organized_Review_Data", index=False)

    print(f"Results saved to {output_filepath}")

# --- Main execution ---
labelata_review_path = "/mnt/hdda/murong/22k/results/Review_File_20250201.xlsx"
output_excel_path = "/mnt/hdda/murong/22k/results/Review_File_20250201_counted.xlsx" # Changed output filename for sorted results

print(f"Reading Labelata review from: {labelata_review_path}")
labelata_review = read_xlsx_to_dict(labelata_review_path)

print("Organizing data by reliability and structure...")
organized_results = organize_by_reliability_and_structure(labelata_review)

print("Calculating reliability counts and sorting...")
df_reliability_counts = calculate_reliability_counts(organized_results)

print(f"Saving sorted results to Excel: {output_excel_path}")
save_results_to_excel(organized_results, df_reliability_counts, output_excel_path)

print("Script finished.")