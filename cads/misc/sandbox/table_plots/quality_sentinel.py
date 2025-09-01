from openpyxl import load_workbook
import os
import pandas as pd
from openpyxl.styles import PatternFill

def read_xlsx_to_dict(file_path):
    data = {}

    wb = load_workbook(filename=file_path, read_only=True)
    sheet = wb.active

    headers = [cell.value for cell in next(sheet.rows)]
    
    image_id_idx = headers.index("Image ID")
    label_idx = headers.index("Label")
    reliable_idx = headers.index("Relaible (Y/N)")
    comment_idx = headers.index('Comment')

    for row in sheet.iter_rows(min_row=2):  # skip header
        image_id = row[image_id_idx].value
        label = row[label_idx].value
        reliable = str(row[reliable_idx].value).strip().upper()

        # process reliability
        if reliable == 'NONE':
            comment = row[comment_idx].value
            if 'not present' in comment or 'minimal part present' in comment:
                # limited field of view, ignore any potential FN
                is_reliable = -2
            elif 'not labelled' in comment:
                # unreliable GT
                is_reliable = False
            else:
                is_reliable = True
        elif reliable == 'TBC':
            is_reliable = -2
        elif reliable == 'N':
            is_reliable = False
        elif reliable == 'Y':
            is_reliable = True

        if image_id not in data:
            data[image_id] = {}
        data[image_id][label] = is_reliable

    return data

def analyze_agreement(df_gt, labelata_review):
    # Get all structure columns
    structure_columns = [col for col in df_gt.columns if col not in ['CT_Sample']]
    
    # Initialize metrics
    metrics = {
        'structure': [],
        'total_unreliable': [],
        'model_hits': [],
        'hit_rate': []
    }
    
    # For each structure
    for structure in structure_columns:
        # Get expert unreliable cases
        expert_unreliable = []
        for img_id in df_gt['CT_Sample']:
            if img_id in labelata_review and structure in labelata_review[img_id]:
                if labelata_review[img_id][structure] is False:
                    expert_unreliable.append(img_id)
        
        if not expert_unreliable:
            continue
            
        # Sort by model scores (ascending - lower score means less reliable)
        sorted_scores = df_gt.sort_values(by=structure)[['CT_Sample', structure]]
        
        # Get model's top N unreliable cases (N = number of expert unreliable cases)
        model_unreliable = sorted_scores['CT_Sample'].head(len(expert_unreliable)).tolist()
        
        # Calculate hits
        hits = len(set(model_unreliable) & set(expert_unreliable))
        hit_rate = hits / len(expert_unreliable) if expert_unreliable else 0
        
        # Store metrics
        metrics['structure'].append(structure)
        metrics['total_unreliable'].append(len(expert_unreliable))
        metrics['model_hits'].append(hits)
        metrics['hit_rate'].append(hit_rate)
        
        print(f"Structure: {structure}")
        print(f"Expert unreliable cases: {len(expert_unreliable)}")
        print(f"Model hits: {hits}")
        print(f"Hit rate: {hit_rate:.2%}")
        print("-" * 50)
    
    return pd.DataFrame(metrics)

def create_highlighted_excel(df_gt, labelata_review, output_path):
    # Create a copy of the DataFrame
    df_highlight = df_gt.copy()
    
    # Create Excel writer
    writer = pd.ExcelWriter(output_path, engine='openpyxl')
    df_highlight.to_excel(writer, index=False, sheet_name='Sheet1')
    
    # Get the worksheet
    worksheet = writer.sheets['Sheet1']
    
    # Define yellow fill
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    
    # Get column indices
    col_indices = {col: idx + 1 for idx, col in enumerate(df_gt.columns)}
    
    # Highlight cells
    for idx, row in df_gt.iterrows():
        img_id = row['CT_Sample']
        if img_id in labelata_review:
            for structure in df_gt.columns:
                if structure != 'CT_Sample' and structure in labelata_review[img_id]:
                    if labelata_review[img_id][structure] is False:
                        cell = worksheet.cell(row=idx + 2, column=col_indices[structure])
                        cell.fill = yellow_fill
    
    # Add average scores row
    last_row = len(df_gt) + 2  # +2 because of header and 1-based indexing
    worksheet.cell(row=last_row, column=1, value="Average Score")
    
    # Calculate and add averages for each structure
    for structure in df_gt.columns:
        if structure != 'CT_Sample':
            # Calculate average of available scores
            avg_score = df_gt[structure].mean()
            col_idx = col_indices[structure]
            worksheet.cell(row=last_row, column=col_idx, value=avg_score)
    
    writer.close()

# Main execution
labelata_review_path = "/mnt/hdda/murong/22k/results/Review File_20250201.xlsx"
labelata_review = read_xlsx_to_dict(labelata_review_path)

output_filename = 'totalseg_cads_test_555'

scores_gt_path = '/mnt/hdda/murong/quality-sentinel/output/results_totalseg_test_555.csv'
checking_path = '/mnt/hdda/murong/quality-sentinel/check_ranking'
output_path = os.path.join(checking_path, f'highlighted_scores_{output_filename}.xlsx')

if not os.path.exists(checking_path):
    os.makedirs(checking_path, exist_ok=True)

# Read model scores
df_gt = pd.read_csv(scores_gt_path)

# Analyze agreement between model and expert review
# metrics_df = analyze_agreement(df_gt, labelata_review)
# metrics_df.to_excel(os.path.join(checking_path, f'unreliable_detection_{output_filename}.xlsx'), index=False)

# Create highlighted Excel file
create_highlighted_excel(df_gt, labelata_review, output_path)
