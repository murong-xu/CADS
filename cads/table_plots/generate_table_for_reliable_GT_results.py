"""
Use Labelata's review to filter out unreliable GT's results in TotalSeg dataset. 
"""

from openpyxl import load_workbook
import os
import pandas as pd

def read_xlsx_to_dict(file_path):
    data = {}

    wb = load_workbook(filename=file_path, read_only=True)
    sheet = wb.active

    headers = [cell.value for cell in next(sheet.rows)]
    
    image_id_idx = headers.index("Image ID")
    label_idx = headers.index("Label")
    reliable_idx = headers.index("Relaible (Y/N)")
    comment_idx = headers.index('Comment')

    for row in sheet.iter_rows(min_row=2):  # skip header
        image_id = row[image_id_idx].value
        label = row[label_idx].value
        reliable = str(row[reliable_idx].value).strip().upper()

        # process reliability
        if reliable == 'NONE':
            comment = row[comment_idx].value
            if 'not present' in comment or 'minimal part present' in comment:
                # limited field of view, ignore any potential FN
                is_reliable = -2
            elif 'not labelled' in comment:
                # unreliable GT
                is_reliable = False
            else:
                is_reliable = True
        elif reliable == 'TBC':
            is_reliable = -2
        elif reliable == 'N':
            is_reliable = False
        elif reliable == 'Y':
            is_reliable = True

        if image_id not in data:
            data[image_id] = {}
        data[image_id][label] = is_reliable

    return data

labelata_review_path = "/mnt/hdda/murong/22k/results/Review File_20250201.xlsx"
labelata_review = read_xlsx_to_dict(labelata_review_path)

# scores_path = '/mnt/hdda/murong/22k/ct_predictions/final_models/scores_labelata/test_0/0037_totalsegmentator'
scores_path = '/mnt/hdda/murong/22k/ct_predictions/baselines/totalseg/metrics_labelata/test_0/0037_totalsegmentator'
# filtered_scores_path = '/mnt/hdda/murong/22k/ct_predictions/final_models/scores_labelata_confirmed_reliable_GT/test_0/0037_totalsegmentator'
filtered_scores_path = '/mnt/hdda/murong/22k/ct_predictions/baselines/totalseg/metrics_labelata_confirmed_reliable_GT/test_0/0037_totalsegmentator'

if not os.path.exists(filtered_scores_path):
    os.makedirs(filtered_scores_path, exist_ok=True)

for filename in os.listdir(scores_path):
    if filename.endswith('.xlsx'):
        file_path = os.path.join(scores_path, filename)
        
        # Read the xlsx file into pandas DataFrame
        df = pd.read_excel(file_path)
        
        # Get image IDs
        image_ids = df['ids'].tolist()
        
        # For each image ID and each structure
        for idx, image_id in enumerate(image_ids):
            if image_id in labelata_review:
                # Get all column names except 'ids'
                structure_columns = [col for col in df.columns if col not in ['ids', 'split', 'background']]
                
                # Check each structure's reliability
                for structure in structure_columns:
                    if structure in labelata_review[image_id]:
                        reliability = labelata_review[image_id][structure]
                        
                        # Update value based on reliability
                        if reliability == -2:
                            df.loc[idx, structure] = -2
                        elif reliability is False:
                            df.loc[idx, structure] = -4
                        # If reliability is True, keep original value
                    else:
                        print(f'No Labelata review found for structure {structure}!')
            else:
                print(f'No Labelata review found for img_id {image_id}!')
        # Save modified DataFrame to new location
        output_file = os.path.join(filtered_scores_path, filename)
        df.to_excel(output_file, index=False)
        print(f"Saved filtered results to {output_file}")

